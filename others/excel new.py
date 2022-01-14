import openpyxl
import json
from collections import OrderedDict
import pprint

with open("lang.json") as f:
    df = json.load(f)

dictlst = {}

for i in df:
    language = i["language"]
    count = i["count"]
    if language not in dictlst.keys():
        dictlst[language] = count
    else:
        dictlst[language] += count

print(dictlst)
wb = openpyxl.load_workbook("json.xlsx")
ws = wb.worksheets[0]
cnt = 1
for j,k in dictlst.items():
    ws.cell(row=cnt, column=1).value = j
    ws.cell(row=cnt, column=2).value = k
    cnt += 1

wb.save("json.xlsx")
